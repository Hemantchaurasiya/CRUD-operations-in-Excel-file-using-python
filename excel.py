import openpyxl
import os

class Excel:
    def __init__(self):
        if os.path.exists("student.xlsx"):
            self.wb=openpyxl.load_workbook("student.xlsx")
            #get the active sheet
            self.sheet=self.wb.active
        else:
            self.wb = openpyxl.Workbook()
            self.wb.save(filename='student.xlsx')
            self.wb=openpyxl.load_workbook("student.xlsx")
            self.sheet=self.wb.active
            self.sheet.cell(row=1,column=1,value="Roll No.")
            self.sheet.cell(row=1,column=2,value="Name")
            self.sheet.cell(row=1,column=3,value="Email")
            self.sheet.cell(row=1,column=4,value="Phone")
            self.sheet.cell(row=1,column=5,value="Year")
            self.sheet.cell(row=1,column=6,value="Department")
            self.wb.save("student.xlsx")

    def insert_user(self,userid,username,email,phone,year,department):
        rows = self.sheet.max_row
        self.sheet.cell(row=rows+1,column=1,value=userid)
        self.sheet.cell(row=rows+1,column=2,value=username)
        self.sheet.cell(row=rows+1,column=3,value=email)
        self.sheet.cell(row=rows+1,column=4,value=phone)
        self.sheet.cell(row=rows+1,column=5,value=year)
        self.sheet.cell(row=rows+1,column=6,value=department)
        self.wb.save("student.xlsx")

    def fetch_all_user(self):
        rows = self.sheet.max_row
        for i in range(2,rows+1):
            print("Enroll No. : ",self.sheet.cell(row=i,column=1).value)
            print("Name : ",self.sheet.cell(row=i,column=2).value)
            print("Email : ",self.sheet.cell(row=i,column=3).value)
            print("Phone : ",self.sheet.cell(row=i,column=4).value)
            print("Year : ",self.sheet.cell(row=i,column=5).value)
            print("Department : ",self.sheet.cell(row=i,column=6).value)
            print()
            print()

    def fetch_user(self,user_id):
        rows = self.sheet.max_row
        for i in range(2,rows+1):
            if self.sheet.cell(row=i,column=1).value == user_id:
                print()
                print()
                print("Enroll No. : ",self.sheet.cell(row=i,column=1).value)
                print("Name : ",self.sheet.cell(row=i,column=2).value)
                print("Email : ",self.sheet.cell(row=i,column=3).value)
                print("Phone : ",self.sheet.cell(row=i,column=4).value)
                print("Year : ",self.sheet.cell(row=i,column=5).value)
                print("Department : ",self.sheet.cell(row=i,column=6).value)
                print()
                print()
                
    def update_user(self,user_id,username,email,phone,year,department):
        rows = self.sheet.max_row
        for i in range(2,rows+1):
            if self.sheet.cell(row=i,column=1).value == user_id:
                self.sheet.cell(row=i,column=2,value=username)
                self.sheet.cell(row=i,column=3,value=email)
                self.sheet.cell(row=i,column=4,value=phone)
                self.sheet.cell(row=i,column=5,value=year)
                self.sheet.cell(row=i,column=6,value=department)
                self.wb.save("student.xlsx")

    def delete_user(self,user_id):
        rows = self.sheet.max_row
        for i in range(2,rows+1):
            if self.sheet.cell(row=i,column=1).value == user_id:
                self.sheet.delete_rows(i,1)
                self.wb.save("student.xlsx")


